"""
Parsers para conciliación bancaria RALUMIN.
- parse_estado_cuenta_bcp: lee el PDF del estado de cuenta corriente BCP.
- parse_excel_contador: lee el Excel de conciliación del contador.
Ambos devuelven listas de diccionarios homogéneos.
"""
import re
import io
import pdfplumber
import openpyxl

# ---------------------------------------------------------------------------
# Utilidades de montos
# ---------------------------------------------------------------------------
# Acepta: "1,234.56"  "123.45"  ".35"  con guion final opcional (= cargo/salida)
_MONTO_RE = re.compile(r'^-?(?:[\d,]*\.\d{2})-?$')
_FECHA_RE = re.compile(r'^\d{2}-\d{2}$')


def _es_monto(token: str) -> bool:
    return bool(_MONTO_RE.match(token))


def _a_float(token: str) -> float:
    neg = token.endswith('-')
    token = token.rstrip('-').replace(',', '')
    val = float(token)
    return round(-val if neg else val, 2)


# ---------------------------------------------------------------------------
# PDF del banco (BCP)
# ---------------------------------------------------------------------------
def _limpiar_bytes_pdf(data: bytes) -> bytes:
    """El PDF del BCP a veces llega con basura ($BOP$ ... $EOP$) antes/después
    de los marcadores reales. Recortamos al PDF válido."""
    inicio = data.find(b"%PDF")
    fin = data.rfind(b"%%EOF")
    if inicio == -1:
        raise ValueError("El archivo no contiene un PDF válido (no se encontró '%PDF').")
    if fin == -1:
        return data[inicio:]
    return data[inicio: fin + 5]


def parse_estado_cuenta_bcp(file_bytes: bytes):
    """
    Devuelve (movimientos, resumen).

    movimientos: lista de dicts con
        fecha (str dd-mm), descripcion (str), monto (float; + abono, - cargo),
        saldo (float), tipo_bcp (str código contable BCP si se detecta),
        es_abono (bool).
    resumen: dict con saldo_inicial, saldo_final, total_abonos, total_cargos,
             moneda ('SOLES'|'DOLARES'|'') y cuenta (str) o None si no se halla.
    """
    pdf_bytes = _limpiar_bytes_pdf(file_bytes)
    movimientos = []
    resumen = None
    moneda = ""
    cuenta = ""

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            texto = page.extract_text() or ""

            # detectar moneda y número de cuenta (aparecen en el encabezado)
            if not moneda:
                up = texto.upper()
                if "SOLES" in up:
                    moneda = "SOLES"
                elif "DOLARES" in up or "DÓLARES" in up:
                    moneda = "DOLARES"
            if not cuenta:
                mcta = re.search(r'\b(\d{3}-\d{7}-\d-\d{2})\b', texto)
                if mcta:
                    cuenta = mcta.group(1)

            for linea in texto.split("\n"):
                tokens = linea.split()

                # --- fila de RESUMEN DEL MES (9 números, sin fecha) ---
                if resumen is None and len(tokens) == 9 and all(_es_monto(t) for t in tokens):
                    try:
                        resumen = {
                            "saldo_inicial": _a_float(tokens[0]),
                            "abonos_efectivo": _a_float(tokens[1]),
                            "abonos_otros": _a_float(tokens[2]),
                            "cargos_cheques": _a_float(tokens[3]),
                            "cargos_otros": _a_float(tokens[4]),
                            "saldo_final": _a_float(tokens[7]),
                        }
                        resumen["total_abonos"] = round(
                            resumen["abonos_efectivo"] + resumen["abonos_otros"], 2)
                        resumen["total_cargos"] = round(
                            -(resumen["cargos_cheques"] + resumen["cargos_otros"]), 2)
                    except Exception:
                        resumen = None
                    continue

                # --- líneas de movimiento (empiezan con fecha dd-mm) ---
                if len(tokens) < 4 or not _FECHA_RE.match(tokens[0]):
                    continue
                idx_montos = [i for i, t in enumerate(tokens) if _es_monto(t)]
                # se requieren al menos 2 montos y el último debe ser el saldo (final de línea)
                if len(idx_montos) < 2 or idx_montos[-1] != len(tokens) - 1:
                    continue

                i_saldo = idx_montos[-1]
                i_monto = idx_montos[-2]
                monto = _a_float(tokens[i_monto])
                saldo = _a_float(tokens[i_saldo])
                descripcion = " ".join(tokens[1:i_monto])

                # intentar aislar el código contable BCP (4 dígitos antes del monto)
                tipo_bcp = ""
                if i_monto >= 1 and re.match(r'^\d{4}$', tokens[i_monto - 1]):
                    tipo_bcp = tokens[i_monto - 1]

                movimientos.append({
                    "fecha": tokens[0],
                    "descripcion": descripcion,
                    "monto": monto,
                    "saldo": saldo,
                    "tipo_bcp": tipo_bcp,
                    "es_abono": monto > 0,
                })

    if resumen is not None:
        resumen["moneda"] = moneda
        resumen["cuenta"] = cuenta

    return movimientos, resumen


def validar_estado_cuenta(movimientos, resumen):
    """Verifica que los movimientos cuadren con el resumen del mes.
    Devuelve dict con el chequeo (útil para mostrar confianza al usuario)."""
    suma_ab = round(sum(m["monto"] for m in movimientos if m["monto"] > 0), 2)
    suma_ca = round(sum(m["monto"] for m in movimientos if m["monto"] < 0), 2)
    out = {
        "n_movimientos": len(movimientos),
        "suma_abonos": suma_ab,
        "suma_cargos": suma_ca,
    }
    if resumen:
        out["abonos_ok"] = abs(suma_ab - resumen["total_abonos"]) < 0.05
        # total_cargos del resumen ya es negativo; suma_ca también es negativo
        out["cargos_ok"] = abs(suma_ca - resumen["total_cargos"]) < 0.05
        # encadenar saldos
        s = resumen["saldo_inicial"]
        encadena = True
        for m in movimientos:
            s = round(s + m["monto"], 2)
            if abs(s - m["saldo"]) > 0.02:
                encadena = False
                break
        out["saldo_final_calculado"] = s
        out["saldos_encadenan"] = encadena
        out["cuadre_total"] = out["abonos_ok"] and out["cargos_ok"] and encadena
    return out


# ---------------------------------------------------------------------------
# Excel del contador
# ---------------------------------------------------------------------------
# Encabezados esperados en la fila de títulos (col B en adelante)
_COL = {
    "mes": 2,        # B  M.
    "subdiario": 3,  # C  S/D
    "asiento": 4,    # D  ASIE
    "concepto": 5,   # E  CONCEPTO
    "fecha_doc": 6,  # F  FECHA D.
    "tipo_doc": 7,   # G  D.
    "numero_doc": 8, # H  NUMERO
    "debe": 9,       # I  DEBE
    "haber": 10,     # J  HABER
    "saldo": 11,     # K  SALDO
    "entidad": 13,   # M  ENTIDAD (RUC)
    "razon": 14,     # N  RAZON SOCIAL
}


def parse_excel_contador(file_bytes: bytes):
    """
    Devuelve (asientos, saldo_inicial).

    asientos: lista de dicts con
        asiento (int|str), subdiario, concepto, numero_doc, debe, haber,
        monto (float; debe - haber  => + entrada, - salida),
        ruc, razon_social, fila_excel (int).
    saldo_inicial: float tomado de la fila 'SALDO AL MES DE ...'.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    # localizar fila de encabezado (la que contiene 'ASIE' y 'CONCEPTO')
    fila_header = None
    for r in range(1, min(ws.max_row, 40) + 1):
        valores = [str(ws.cell(r, c).value or "").strip().upper() for c in range(2, 15)]
        if "ASIE" in valores and "CONCEPTO" in valores:
            fila_header = r
            break
    if fila_header is None:
        fila_header = 8  # valor por defecto observado

    asientos = []
    saldo_inicial = 0.0

    for r in range(fila_header + 1, ws.max_row + 1):
        asiento = ws.cell(r, _COL["asiento"]).value
        concepto = ws.cell(r, _COL["concepto"]).value
        concepto = (str(concepto).strip() if concepto is not None else "")

        # fila de saldo inicial
        if "SALDO AL MES" in concepto.upper():
            sal = ws.cell(r, _COL["saldo"]).value
            if isinstance(sal, (int, float)):
                saldo_inicial = round(float(sal), 2)
            continue

        # fila vacía / sin datos útiles
        if asiento is None and not concepto:
            continue

        debe = ws.cell(r, _COL["debe"]).value
        haber = ws.cell(r, _COL["haber"]).value
        debe = float(debe) if isinstance(debe, (int, float)) else 0.0
        haber = float(haber) if isinstance(haber, (int, float)) else 0.0

        razon = ws.cell(r, _COL["razon"]).value
        ruc = ws.cell(r, _COL["entidad"]).value
        numdoc = ws.cell(r, _COL["numero_doc"]).value
        sub = ws.cell(r, _COL["subdiario"]).value

        asientos.append({
            "asiento": asiento,
            "subdiario": (str(sub).strip() if sub is not None else ""),
            "concepto": concepto,
            "numero_doc": (str(numdoc).strip() if numdoc is not None else ""),
            "debe": round(debe, 2),
            "haber": round(haber, 2),
            "monto": round(debe - haber, 2),
            "ruc": (str(ruc).strip() if ruc is not None else ""),
            "razon_social": (str(razon).strip() if razon is not None else ""),
            "fila_excel": r,
        })

    return asientos, saldo_inicial
